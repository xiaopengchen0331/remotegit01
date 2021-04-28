import openpyxl

wb = openpyxl.load_workbook('t01.xlsx')
print(wb.sheetnames)
mySheet = wb.create_sheet('cxp')
for sheet in wb:
    print(sheet.title)
ws = wb.active
print(ws)
print(ws['A1'])
print(ws['A1'].value)
c = ws['B2']
print(c.row, c.column, c.value)
print(c.coordinate)
#wb.save('t01.xlsx')
#wb.save('t02.xlsx')